"""
Workbook class for xls2latex
Created on Mon Mar 23 12:47:40 2020

@author: ppenguin
"""

import openpyxl
from xls2latex import xlcellfun

class xlTableTeX(object):

    def __init__(self, worksheet, textcharwidth=80, caption=None, label=None, colwidths=None, vfix=None, smalltext=False):

        self.sheet = worksheet
        self.minco = (self.sheet.min_row, self.sheet.min_column)
        self.maxco = (self.sheet.max_row, self.sheet.max_column)
        self.numcols = self.sheet.max_column - self.sheet.min_column + 1
        self.numrows = self.sheet.max_row - self.sheet.min_row + 1
        self.textcharwidth = textcharwidth
        self.caption = caption
        self.label = label
        self.setcolwidths(colwidths)
        self.vfix = vfix
        self.smalltext = smalltext
        # difference between paragraph and normal columns https://www.overleaf.com/learn/latex/tables
        self.pvalignmap = {None: 'p', 'bottom': 'b', 'center': 'm', 'distributed': 'm', 'justify': 'm', 'top': 'p'}
        self.phalignmap = {None: '\\raggedright', 'left': '\\raggedright', 'centerContinuous': '\\centering', 'center': '\\centering', 'distributed': '\\centering', 'fill': '\\centering', 'justify': '\\centering', 'right': '\\raggedleft', 'general': '\\raggedright'}
        self.halignmap = {None: 'l', 'left': 'l', 'centerContinuous': 'c', 'center': 'c', 'distributed': 'c', 'fill': 'c', 'justify': 'c', 'right': 'r', 'general': 'l'}
        self.setdmergedrngs()
        self.tbody = ""
        self.texout = ""


    # pre-make a dict of the start of merged ranges so that we only need to scan once
    def setdmergedrngs(self):
        self.dmergedranges = {}
        for rng in self.sheet.merged_cells.ranges:
            self.dmergedranges[(rng.min_row, rng.min_col)] = rng

    # returns the colspan of a merged cell if we are at the start (left corner) of the range
    # if the cell is not merged, the colspan is 1
    def getcolspan(self, r, c):
        if (r, c) in self.dmergedranges:
            return self.dmergedranges[(r ,c)].max_col - self.dmergedranges[(r ,c)].min_col + 1
        else:
            return 1

    # returns the rowspan if we are at the start (top corner) of the range
    def getrowspan(self, r, c):
        if (r, c) in self.dmergedranges:
            return self.dmergedranges[(r ,c)].max_row - self.dmergedranges[(r ,c)].min_row + 1
        else:
            return 1

    ### obsolete, we just do the book-keeping in the loop
    # check whether the current coordinates are within a merged range
    # not so efficient since we execute this loop frequently
    def inmerged(self, r, c):
        # the check range is the current coordinate (1 cell)
        cr = openpyxl.worksheet.cell_range.CellRange(min_col=c, min_row=r, max_col=c, max_row=r)
        ret = False
        for mr in self.sheet.merged_cells.ranges:
            ret = cr.issubset(mr)
            if ret:
                return ret

        return ret


    # check whether te cell with coordinates (r,c) is in a vertically merged range (except last row)
    def invmerged(self, r, c):
        # the check range is the current coordinate (1 cell)
        cr = openpyxl.worksheet.cell_range.CellRange(min_col=c, min_row=r, max_col=c, max_row=r)
        ret = False
        for mr in self.sheet.merged_cells.ranges:
            if cr.issubset(mr): # we have found the merged range of which the current (c,r) is a part
                if mr.max_row > mr.min_row: # vertical range
                    return (r >= mr.min_row and r < mr.max_row)  # we are in mr, except bottom

        return ret


    def getcellhalign(self, c):
        if c.alignment.horizontal is not None:
            if  c.alignment.horizontal[0] == 'g':
                return 'l' # return left for general
            else:
                return c.alignment.horizontal[0]
        else:
            return 'l'


    def getcellvalign(self, c):
        if c.alignment.vertical is not None:
            return c.alignment.vertical[0]
        else:
            return 't'


    # we could make this more advanced?
    def getbordersym(self, b):
        if b is None or b.style is None:
            return ""
        else:
            return "|"


    def gencmidrule(self, r):
        currstyle = None
        ss = None
        ruledef = ""
        for c in range(self.minco[1], self.maxco[1] + 1):
            if (r, c) in self.sheet._cells.keys():
                style = self.sheet._cells[(r, c)].border.bottom.style
                # we have to ignore the bottom border if we are in a rowspan, unless at the bottom
                if self.invmerged(r, c):
                    style = None
            else:
                style = None
            if currstyle != style or ss == None or c == self.maxco[1]: # we have a style change or we just started or are at the end
                if ss is not None: # flush the current style to the cmidruledef
                    if currstyle == 'double':
                        ruledef += "\\cmidrule{%(ss)d-%(c)d}\\morecmidrules\\cmidrule{%(ss)d-%(c)d}"%vars()
                    elif currstyle is not None:
                        ruledef += "\\cmidrule{%(ss)d-%(c)d}"%vars()
                currstyle = style
                ss = c  # style start is current column
        return ruledef


    def applytexformat(self, cv, oc):
        tex = cv
        if oc.font.b:
            tex = "\\textbf{%s}"%tex
        if oc.font.i:
            tex = "\\textit{%s}"%tex
        return tex


    def setcolwidths(self, s):
        l = []
        if s is not None and len(s) > 0:
            for p in s.split(","):
                try:
                    f = float(p)
                    l.append("{%0.2f\\textwidth}"%f)
                except:
                    l.append("{%s}"%p)
        self.colwidths = l

    def texescape(self, s):
        # for now escape all % with \
        # (other escapes here too?)
        return s.replace(r"%", r"\%")

    # get a complete string e.g. for a table definition
    # self-specified list of aligns not yet implemented
    #def getcolwidthsstr(self, lhalign=None, lvalign=None):
    #    #if len(lalign) <> len(self.colwidths):
    #    l = []
    #    for s in self.colwidths:
    #        # s has a value of the form {0.3\textwidth}
    #       # including align this has to be >{\centering}p{0.3\textwidth}
    #        if self.colwidths.index(s) <= len(lhalign): # get info for this col
    #
    #    return "".join(l)


    # https://stackoverflow.com/questions/1110961/how-can-i-restrict-the-size-of-my-multicolumn-cells-in-a-longtable
    # get a complete multicolumn format string with column width, alignment and borders
    def getmcformatstr(self, c=None, al=None, bo=None, colcount=1):

        # this is the 1-based column number, if not given we cannot get column defaults and return the simple representation
        # if we have a colspan, ignpre column widths specified and return simple representation
        ret = ''

        if c is None or len(self.colwidths) == 0 or c > len(self.colwidths) or colcount > 1:
            if al is not None:
                alstr = self.halignmap[al.horizontal]
            else:
                alstr = 'l'
            if bo is not None:
                ret = self.getbordersym(bo.left) + alstr + self.getbordersym(bo.right)
            else:
                ret = alstr

            return ret

        # if we come here, we have a column number and column width info
        if al is not None:
            halstr = ">{%s}"%self.phalignmap[al.horizontal]
            valstr = self.pvalignmap[al.vertical]
        else:
            halstr = ">{\\raggedright}"
            valstr = "p"

        if bo is not None:
            lb = self.getbordersym(bo.left)
            rb = self.getbordersym(bo.right)
        else:
            lb = ''
            rb = ''

        ret = halstr + lb + valstr + rb + self.colwidths[c-1]

        return ret


    def genTex(self, width=80, caption=None, label=None, colwidths=None, vfix=None):

        # always use multicol and multirow for flexibility reasons

        # apparently the _cells property shows merged cells incorrectly
        # (the start is 1 col/row too late, and the length one less)
        # so we use

        # make body
        tbody = "\\toprule\n"
        endrowspan = {}
        inrowspan = {}

        for ro in range(self.minco[0], self.maxco[0] + 1):
            endcolspan = 0
            incolspan = False
            for co in range(self.minco[1], self.maxco[1] + 1):

                if not co in inrowspan.keys():
                    inrowspan[co] = False
                if not co in endrowspan.keys():
                    endrowspan[co] = 0

                incolspan = (incolspan and co <= endcolspan)
                inrowspan[co] = (inrowspan[co] and ro <= endrowspan[co])

                cs = self.getcolspan(ro, co)
                if cs > 1:
                    endcolspan = co + cs - 1
                    incolspan = True

                rs = self.getrowspan(ro, co)
                if rs > 1:
                    endrowspan[co] = ro + rs - 1
                    inrowspan[co] = True

                mctext = ""
                mrtext = ""
                celltext = ""
                # if we are still in the table range, it is possible that there are undefined cells
                # so we check for them and make an empty cell if there is no definition
                if (ro, co) in self.sheet._cells.keys():
                    ocell = self.sheet._cells[(ro, co)]
                    cval = xlcellfun.reformat_cellval(ocell)
                    cval = self.applytexformat(cval, ocell)
                    al = ocell.alignment
                    bo = ocell.border
                else:   # define some defaults so we can execute the normal code
                    ocell = None
                    cval = ""
                    al = None
                    bo = None

                # if we are within a merged cell, but not at the start, we skip the content
                if rs > 1:  # we are at the start of a merged row
                    valign = self.getcellvalign(ocell) if ocell is not None else valign
                    vmove = "%0.2fex"%(-rs/2.0) if vfix is None else vfix
                    mrtext = "\\multirow[%s]{%d}{*}[%s]"%(valign, rs, vmove)

                # For every multicolumn, if we specify a fixed width for the column in the table env,
                # we also **must** specify it for the cell, otherwise it's ignored...
                if cs > 1 or rs > 1 or (not inrowspan[co] and not incolspan): # start of colspan or a normal cell (we still use a multicol for easier alignment and borders)
                    mctext += "\\multicolumn{%d}{%s}{"%(cs, self.getmcformatstr(co, al, bo, colcount=cs))
                elif inrowspan[co] and not incolspan: # here we need an empty cell
                    mrtext = " "
                    # actually we need an empty cell with a border, so:
                    mctext += "\\multicolumn{%d}{%s}{}"%(cs, self.getmcformatstr(co, al, bo))

                # multicolumn needs to go before multirow (apparently, otherwise errors)
                if len(mctext)>0:
                    if mrtext == "":
                        mrtext += "%s}"%cval
                    elif mrtext != " ":
                        mrtext += "{%s}}"%cval

                celltext = mctext + mrtext

                if len(celltext) > 0 and co < self.maxco[1] and not incolspan:  # the last condition is for colspans that are full-width
                    celltext += " & "

                if co == self.maxco[1]:
                    celltext += " \\\\ \n"
                    # handle \cmidrule here! (we are at the end of a row and have all the info for the cells in this row)
                    if ro < self.maxco[0]:  # skip for last line, since we do a bottomrule
                        cmr = self.gencmidrule(ro)
                        if len(cmr) > 0:
                            celltext += cmr + " \n"

                tbody += self.texescape(celltext)


        # define the environment (for now whe just choose default left alignment, since we specifically align each cell with multicol)
        comment = """% For vertical borders to work with booktabs, you **need** to put the following in the preamble:
% \\aboverulesep=0ex
% \\belowrulesep=0ex
% uncramp the table (due to the lost vertical space in cmidrule etc.
% \\renewcommand{\\arraystretch}{1.4} % choose this factor larger than 1.0 \n\n"""

        if self.smalltext:
            smallbeg = "{\small\\tabcolsep=3.5pt  % hold it local \n"
            smallend = "}% small text \n"
        else:
            smallbeg = ""
            smallend = ""

        # if colwidths is None:
        tenvbeg = "\\begin{longtable}{%s} \n"%("l"*self.numcols)
        # else:
        #    tenvbeg = "\\begin{longtable}{%s} \n"%(self.getcolwidthsstr(colwidths)) # this is not sufficient if all cells are multicolumn! It has to be specified per cell!
        tenvend = "\\bottomrule \n\\end{longtable} \n"

        caplab = ""
        if self.caption is not None:
            caplab += "\\caption{%s}"%self.caption
        if self.label is not None:
            caplab += "\\label{%s}"%self.label
        if len(caplab) > 0:
            caplab += "\\\\ \n"

        self.texout = comment + smallbeg + tenvbeg + caplab + tbody + tenvend + smallend
