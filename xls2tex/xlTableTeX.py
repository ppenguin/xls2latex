"""
Workbook class for xls2tex
Created on Mon Mar 23 12:47:40 2020

@author: ppenguin
"""

import openpyxl
from . import xlcellfun

class xlTableTeX(object):
    
    def __init__(self, worksheet):
        
        self.sheet = worksheet
        self.minco = (self.sheet.min_row, self.sheet.min_column)
        self.maxco = (self.sheet.max_row, self.sheet.max_column)
        self.numcols = self.sheet.max_column - self.sheet.min_column + 1
        self.numrows = self.sheet.max_row - self.sheet.min_row + 1
        self.setdmergedrngs()
        self.tbody = ""
        self.texout = ""


    # pre-make a dict of the start of merged ranges so that we only need to scan once
    def setdmergedrngs(self):
        self.dmergedranges = {}
        for rng in self.sheet.merged_cells.ranges:
            self.dmergedranges[(rng.min_row, rng.min_col)] = rng
            
    # returns the colspan of a merged cell if we are at the start (left corner) of the range
    # if the cell is not merged, the colspan is 1
    def getcolspan(self, r, c):
        if (r, c) in self.dmergedranges:
            return self.dmergedranges[(r ,c)].max_col - self.dmergedranges[(r ,c)].min_col + 1
        else:
            return 1
    
    # returns the rowspan if we are at the start (top corner) of the range
    def getrowspan(self, r, c):
        if (r, c) in self.dmergedranges:
            return self.dmergedranges[(r ,c)].max_row - self.dmergedranges[(r ,c)].min_row + 1
        else:
            return 1   
    
    ### obsolete, we just do the book-keeping in the loop
    # check whether the current coordinates are within a merged range
    # not so efficient since we execute this loop frequently
    def inmerged(self, r, c):
        # the check range is the current coordinate (1 cell)
        cr = openpyxl.worksheet.cell_range.CellRange(min_col=c, min_row=r, max_col=c, max_row=r)
        ret = False
        for mr in self.sheet.merged_cells.ranges:
            ret = cr.issubset(mr)
            if ret:
                return ret
        
        return ret

    
    def getcellhalign(self, c):
        if c.alignment.horizontal is not None:
            if  c.alignment.horizontal[0] == 'g':
                return 'l' # return left for general
            else:
                return c.alignment.horizontal[0]
        else:
            return 'l'


    def getcellvalign(self, c):
        if c.alignment.vertical is not None:
            return c.alignment.vertical[0]
        else:
            return 't'
        
        
    # we could make this more advanced?
    def getbordersym(self, b):
        if b is None or b.style is None:
            return ""
        else:
            return "|" 
        
             
    def gencmidrule(self, r):
        currstyle = None
        ss = None
        ruledef = ""
        for c in range(self.minco[1], self.maxco[1] + 1):
            style = self.sheet._cells[(r, c)].border.bottom.style
            if currstyle != style or ss == None or c == self.maxco[1]: # we have a style change or we just started or are at the end
                if ss is not None: # flush the current style to the cmidruledef
                    if currstyle == 'double':
                        ruledef += "\\cmidrule{%(ss)d-%(c)d}\\morecmidrules\\cmidrule{%(ss)d-%(c)d}"%vars()
                    elif currstyle is not None:
                        ruledef += "\\cmidrule{%(ss)d-%(c)d}"%vars()
                currstyle = style
                ss = c  # style start is current column
        return ruledef 


    def applytexformat(self, cv, oc):
        tex = cv
        if oc.font.b:
            tex = "\\textbf{%s}"%tex
        if oc.font.i:
            tex = "\\textit{%s}"%tex
        return tex
            
        

    def genTex(self, width=80, caption=None, label=None):
        
        # always use multicol and multirow for flexibility reasons
        
        # apparently the _cells property shows merged cells incorrectly
        # (the start is 1 col/row too late, and the length one less)
        # so we use 
        
        # make body
        tbody = "\\toprule\n"
        endrowspan = {}
        inrowspan = {}
        for ro in range(self.minco[0], self.maxco[0] + 1):
            endcolspan = 0
            incolspan = False
            for co in range(self.minco[1], self.maxco[1] + 1):
                
                if not co in inrowspan.keys():
                    inrowspan[co] = False
                if not co in endrowspan.keys():    
                    endrowspan[co] = 0
                
                incolspan = (incolspan and co <= endcolspan)
                inrowspan[co] = (inrowspan[co] and ro <= endrowspan[co])
                
                cs = self.getcolspan(ro, co)
                if cs > 1:
                    endcolspan = co + cs - 1
                    incolspan = True
                
                rs = self.getrowspan(ro, co)
                if rs > 1:
                    endrowspan[co] = ro + rs - 1
                    inrowspan[co] = True
                
                mctext = ""
                mrtext = ""
                celltext = ""
                ocell = self.sheet._cells[(ro, co)]
                cval = xlcellfun.reformat_cellval(ocell)
                cval = self.applytexformat(cval, ocell)
                # if we are within a merged cell, but not at the start, we skip the content
                if rs > 1:  # we are at the start of a merged row
                    valign = self.getcellvalign(ocell)
                    vmove = "%0.2fex"%(-rs/1.4)
                    mrtext = "\\multirow[%s]{%d}{*}[%s]"%(valign, rs, vmove)
                
                # get some cell layout info
                halign = self.getcellhalign(ocell)                  
                lb = self.getbordersym(ocell.border.left)
                rb = self.getbordersym(ocell.border.right)
                
                if cs > 1 or rs > 1 or (not inrowspan[co] and not incolspan): # start of colspan or a normal cell (we still use a multicol for easier alignment and borders)
                    mctext += "\\multicolumn{%d}{%s%s%s}{"%(cs, lb, halign, rb)
                elif inrowspan[co] and not incolspan: # here we need an empty cell
                    mrtext = " "
                    # actually we need an empty cell with a border, so:
                    mctext += "\\multicolumn{%d}{%s%s%s}{}"%(cs, lb, halign, rb)
                
                # multicolumn needs to go before multirow (apparently, otherwise errors)
                if len(mctext)>0:
                    if mrtext == "":
                        mrtext += "%s}"%cval
                    elif mrtext != " ":
                        mrtext += "{%s}}"%cval

                celltext = mctext + mrtext

                if len(celltext) > 0 and co < self.maxco[1]:
                    celltext += " & "

                if co == self.maxco[1]:
                    celltext += " \\\\ \n"
                    # handle \cmidrule here! (we are at the end of a row and have all the info for the cells in this row)
                    if ro < self.maxco[0]:  # skip for last line, since we do a bottomrule
                        cmr = self.gencmidrule(ro)
                        if len(cmr) > 0:
                            celltext += cmr + " \n"
                
                tbody += celltext
        
            
        # define the environment (for now whe just choose default left alignment, since we specifically align each cell with multicol)
        comment = """% For vertical borders to work with booktabs, you **need** to put the following in the preamble:
% \\aboverulesep=0ex
% \\belowrulesep=0ex
% uncramp the table (due to the lost vertical space in cmidrule etc.
% \\renewcommand{\\arraystretch}{1.4} % choose this factor larger than 1.0 \n\n"""
                     
        tenvbeg = "\\begin{longtable}{%s} \n"%("l"*self.numcols)
        tenvend = "\\bottomrule \n\\end{longtable} \n"
        
        caplab = ""
        if caption is not None:
            caplab += "\\caption{%s}"%caption
        if label is not None:
            caplab += "\\label{%s}"%label
        if len(caplab) > 0:
            caplab += "\\\\ \n"
        
        self.texout = comment + tenvbeg + caplab + tbody + tenvend
        